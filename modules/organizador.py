import csv
import re
import shutil
from pathlib import Path

from openpyxl import load_workbook


def ler_cabecalho_planilha(planilha: Path) -> list[str]:
    """Lê só a linha de cabeçalho da planilha (.xlsx ou .csv), sem processar
    as demais linhas - usado pela UI (OrganizadorWidget) pra sugerir, assim
    que o usuário seleciona o arquivo, quais colunas existem de verdade,
    em vez de exigir que ele digite o nome exato de cabeça.
    """
    extensao = planilha.suffix.lower()
    if extensao == '.xlsx':
        wb = load_workbook(filename=str(planilha), read_only=True, data_only=True)
        try:
            cabecalho = next(wb.worksheets[0].iter_rows(values_only=True), None)
        finally:
            wb.close()
    elif extensao == '.csv':
        with open(planilha, newline='', encoding='utf-8-sig') as f:
            primeira_linha = f.readline()
            if not primeira_linha:
                raise ValueError('O arquivo CSV esta vazio.')
            separador = ';' if primeira_linha.count(';') > primeira_linha.count(',') else ','
            f.seek(0)
            cabecalho = next(csv.reader(f, delimiter=separador), None)
    else:
        raise ValueError('Formato de planilha nao suportado. Use .xlsx ou .csv')

    if not cabecalho:
        raise ValueError('A planilha esta vazia ou nao tem linha de cabecalho.')
    return [str(valor).strip() for valor in cabecalho if valor is not None and str(valor).strip()]


class Organizador:
    def __init__(
        self,
        propostas: str,
        pedidos: str,
        destino: str,
        planilha: str,
        col_rc: str,
        col_po: str,
        col_fornecedor: str,
        logger,
    ):
        self.propostas = propostas
        self.pedidos = pedidos
        self.destino = destino
        self.planilha = planilha
        self.col_rc = col_rc
        self.col_po = col_po
        self.col_fornecedor = col_fornecedor
        self.logger = logger

    def log(self, mensagem: str) -> None:
        if callable(self.logger):
            self.logger(mensagem)

    def sanitizar_nome(self, nome: str) -> str:
        caracteres_invalidos = re.compile(r'[\\/:*?"<>|\n\r\t]')
        nome_limpo = caracteres_invalidos.sub(' ', str(nome)).strip()
        nome_limpo = re.sub(r'\s+', ' ', nome_limpo).strip()
        # O Windows remove silenciosamente pontos e espaços no final do nome
        # de uma pasta/arquivo ao criá-la - sem isso, um fornecedor com "."
        # no final (ex: "LTDA .") gera um Path com o ponto, mas a pasta que
        # de fato aparece em disco não tem o ponto, e qualquer cópia
        # posterior pra dentro dela falha com "caminho não encontrado"
        # (WinError 3), mesmo a pasta existindo.
        nome_limpo = nome_limpo.rstrip(' .')
        return nome_limpo if nome_limpo else "SEM_NOME"

    def validar_pasta(self, pasta: Path, nome: str) -> None:
        if not pasta.is_dir():
            raise ValueError(f'A {nome} informada nao existe ou nao e uma pasta: {pasta}')

    def ler_linhas_xlsx(self, planilha: Path, col_rc: str, col_po: str, col_fornecedor: str) -> list:
        wb = load_workbook(filename=str(planilha), read_only=True, data_only=True)
        sheet = wb.worksheets[0]
        linhas_iter = sheet.iter_rows(values_only=True)
        cabecalho = next(linhas_iter, None)
        if not cabecalho:
            raise ValueError('A planilha esta vazia ou nao tem linha de cabecalho.')

        indices = self._mapear_indices_cabecalho(cabecalho, col_rc, col_po, col_fornecedor)
        resultado = []
        for linha in linhas_iter:
            if not linha or all(v is None for v in linha):
                continue
            resultado.append(self._extrair_valores(linha, indices))
        wb.close()
        return resultado

    def ler_linhas_csv(self, planilha: Path, col_rc: str, col_po: str, col_fornecedor: str) -> list:
        with open(planilha, newline='', encoding='utf-8-sig') as f:
            primeira_linha = f.readline()
            if not primeira_linha:
                raise ValueError('O arquivo CSV esta vazio.')
            separador = ';' if primeira_linha.count(';') > primeira_linha.count(',') else ','
            f.seek(0)
            leitor = csv.DictReader(f, delimiter=separador)
            mapa = {}
            alvo = {
                'rc': col_rc.strip().lower(),
                'po': col_po.strip().lower(),
                'fornecedor': col_fornecedor.strip().lower(),
            }
            for campo in leitor.fieldnames or []:
                if campo:
                    campo_norm = str(campo).strip().lower()
                    for chave, nome_alvo in alvo.items():
                        if campo_norm == nome_alvo:
                            mapa[chave] = campo
            faltando = [chave for chave in ('rc', 'po', 'fornecedor') if chave not in mapa]
            if faltando:
                raise ValueError(f'Colunas nao encontradas no cabecalho do CSV: {faltando}')

            resultado = []
            for linha in leitor:
                resultado.append({
                    'rc': str(linha.get(mapa['rc']) or '').strip(),
                    'po': str(linha.get(mapa['po']) or '').strip(),
                    'fornecedor': str(linha.get(mapa['fornecedor']) or '').strip(),
                })
            return resultado

    def ler_linhas_da_planilha(self, planilha: Path, col_rc: str, col_po: str, col_fornecedor: str) -> list:
        extensao = planilha.suffix.lower()
        if extensao == '.xlsx':
            return self.ler_linhas_xlsx(planilha, col_rc, col_po, col_fornecedor)
        elif extensao == '.csv':
            return self.ler_linhas_csv(planilha, col_rc, col_po, col_fornecedor)
        else:
            raise ValueError('Formato de planilha nao suportado. Use .xlsx ou .csv')

    def _mapear_indices_cabecalho(self, cabecalho, col_rc: str, col_po: str, col_fornecedor: str) -> dict:
        indices = {}
        alvo = {
            'rc': col_rc.strip().lower(),
            'po': col_po.strip().lower(),
            'fornecedor': col_fornecedor.strip().lower(),
        }
        for i, valor in enumerate(cabecalho):
            if valor is None:
                continue
            valor_norm = str(valor).strip().lower()
            for chave, nome_alvo in alvo.items():
                if valor_norm == nome_alvo:
                    indices[chave] = i
        faltando = [chave for chave in ('rc', 'po', 'fornecedor') if chave not in indices]
        if faltando:
            raise ValueError(f'Colunas nao encontradas no cabecalho: {faltando}')
        return indices

    def _extrair_valores(self, linha, indices) -> dict:
        def pegar(chave):
            i = indices[chave]
            if i >= len(linha) or linha[i] is None:
                return ''
            val = linha[i]
            # Converte float inteiro (ex: 123456.0 -> 123456) para nao gerar "123456.0"
            if isinstance(val, float) and val == val // 1:
                val = int(val)
            # Converte numero para string sem sufixo .0
            if isinstance(val, float):
                val = str(val).rstrip('0').rstrip('.')
            return str(val).strip()

        return {
            'rc': pegar('rc'),
            'po': pegar('po'),
            'fornecedor': pegar('fornecedor'),
        }

    @staticmethod
    def _extrair_numero(texto: str) -> str:
        """Extrai apenas os dígitos de um texto. Ex: 'PO nº 123456' -> '123456'."""
        numeros = re.findall(r'\d+', texto)
        return ''.join(numeros) if numeros else texto

    def buscar_arquivo_por_codigo(self, arquivos: list, codigo: str) -> list:
        """Busca arquivos pelo código em uma lista pré-indexada (Item 13)."""
        if not codigo:
            return []
        codigo_busca = self._extrair_numero(str(codigo))
        if not codigo_busca:
            return []
        codigo_lower = codigo_busca.lower().strip()
        return [arq for arq in arquivos if codigo_lower in arq.name.lower()]

    def executar(self) -> None:
        pasta_propostas = Path(self.propostas) if self.propostas else None
        pasta_pedidos = Path(self.pedidos) if self.pedidos else None
        pasta_destino = Path(self.destino)
        planilha = Path(self.planilha)

        if not self.planilha:
            raise ValueError('Por favor, selecione a planilha (.xlsx/.csv).')
        if not self.destino:
            raise ValueError('Por favor, selecione a pasta de destino.')
        if not self.propostas and not self.pedidos:
            raise ValueError('Informe pelo menos uma das pastas (Propostas ou Pedidos).')

        if pasta_propostas is not None:
            self.validar_pasta(pasta_propostas, 'pasta_propostas')
        if pasta_pedidos is not None:
            self.validar_pasta(pasta_pedidos, 'pasta_pedidos')
        pasta_destino.mkdir(parents=True, exist_ok=True)

        linhas = self.ler_linhas_da_planilha(planilha, self.col_rc, self.col_po, self.col_fornecedor)
        self.log(f'Total de linhas lidas da planilha: {len(linhas)}')
        if linhas:
            self.log(f'  Colunas mapeadas: RC="{self.col_rc}" | PO="{self.col_po}" | Forn="{self.col_fornecedor}"')
            self.log('  Primeiras linhas:')
            for i, amostra in enumerate(linhas[:3]):
                self.log(f'    [{i+1}] RC="{amostra["rc"]}" | PO="{amostra["po"]}" | Forn="{amostra["fornecedor"]}"')
        self.log('')

        # Item 13: indexa os arquivos das pastas uma única vez (O(n+m) em vez de O(n×m))
        arquivos_propostas = list(pasta_propostas.rglob('*')) if pasta_propostas else []
        arquivos_propostas = [a for a in arquivos_propostas if a.is_file()]
        arquivos_pedidos = list(pasta_pedidos.rglob('*')) if pasta_pedidos else []
        arquivos_pedidos = [a for a in arquivos_pedidos if a.is_file()]

        total_copiados = 0
        total_pastas_criadas = 0
        total_erros = 0
        destinos_copiados = set()
        linhas_sem_nenhum_arquivo = []
        erros_detalhados = []
        linhas_sem_fornecedor = 0

        for index, linha in enumerate(linhas, start=1):
            rc, po, fornecedor = linha['rc'], linha['po'], linha['fornecedor']

            if not fornecedor:
                linhas_sem_fornecedor += 1
                continue

            # SEMPRE cria a pasta do fornecedor, mesmo que nenhum arquivo seja encontrado
            try:
                pasta_fornecedor = pasta_destino / self.sanitizar_nome(fornecedor)
                pasta_nova = not pasta_fornecedor.exists()
                pasta_fornecedor.mkdir(parents=True, exist_ok=True)
                if pasta_nova:
                    total_pastas_criadas += 1
            except Exception as e:
                total_erros += 1
                erro_msg = f'Falha ao criar pasta para fornecedor "{fornecedor}": {str(e)}'
                erros_detalhados.append((linha, erro_msg))
                self.log(f'  ❌ [Linha {index}] {erro_msg}')
                continue

            try:
                propostas_encontradas = (
                    self.buscar_arquivo_por_codigo(arquivos_propostas, rc) if pasta_propostas else []
                )
            except Exception as e:
                propostas_encontradas = []
                erro_msg = f'Erro ao buscar propostas para RC="{rc}": {str(e)}'
                erros_detalhados.append((linha, erro_msg))
                self.log(f'  ⚠️ [Linha {index}] {erro_msg}')

            try:
                pedidos_encontrados = self.buscar_arquivo_por_codigo(arquivos_pedidos, po) if pasta_pedidos else []
            except Exception as e:
                pedidos_encontrados = []
                erro_msg = f'Erro ao buscar pedidos para PO="{po}": {str(e)}'
                erros_detalhados.append((linha, erro_msg))
                self.log(f'  ⚠️ [Linha {index}] {erro_msg}')

            if not propostas_encontradas and not pedidos_encontrados:
                linhas_sem_nenhum_arquivo.append((rc, po, fornecedor))
                self.log(f'  ℹ️ [Linha {index}] Nenhum arquivo encontrado: RC="{rc}" PO="{po}" Forn="{fornecedor}"')
                continue

            rc_sanitizada = self.sanitizar_nome(rc)
            po_sanitizada = self.sanitizar_nome(po)

            if pasta_propostas:
                for origem in propostas_encontradas:
                    try:
                        extensao = origem.suffix
                        if po_sanitizada and po_sanitizada.lower() not in ['nan', 'sem_nome']:
                            nome_destino = f"PROPOSTA - {rc_sanitizada} - {po_sanitizada}{extensao}"
                        else:
                            nome_destino = f"PROPOSTA - {rc_sanitizada}{extensao}"

                        destino = pasta_fornecedor / nome_destino
                        if str(destino).lower() in destinos_copiados:
                            self.log(f'  ℹ️ Aviso: Ignorado (ja copiado): {origem.name} -> {nome_destino}')
                            continue
                        shutil.copy2(origem, destino)
                        destinos_copiados.add(str(destino).lower())
                        self.log(f'  ✅ Proposta: {origem.name} -> {nome_destino}')
                        total_copiados += 1
                    except Exception as e:
                        total_erros += 1
                        erro_msg = f'Falha ao copiar proposta "{origem.name}" para "{fornecedor}": {str(e)}'
                        erros_detalhados.append((linha, erro_msg))
                        self.log(f'  ❌ [Linha {index}] {erro_msg}')

            if pasta_pedidos:
                for origem in pedidos_encontrados:
                    try:
                        destino = pasta_fornecedor / origem.name
                        if str(destino).lower() in destinos_copiados:
                            self.log(f'  ℹ️ Aviso: Ignorado (ja copiado): {origem.name} -> {pasta_fornecedor.name}')
                            continue
                        shutil.copy2(origem, destino)
                        destinos_copiados.add(str(destino).lower())
                        self.log(f'  ✅ Pedido: {origem.name} -> {pasta_fornecedor.name}')
                        total_copiados += 1
                    except Exception as e:
                        total_erros += 1
                        erro_msg = f'Falha ao copiar pedido "{origem.name}" para "{fornecedor}": {str(e)}'
                        erros_detalhados.append((linha, erro_msg))
                        self.log(f'  ❌ [Linha {index}] {erro_msg}')

        self.log('')
        self.log('===== RESUMO DA EXECUCAO =====')
        self.log(f'Arquivos copiados com sucesso: {total_copiados}')
        self.log(f'Pastas de fornecedores criadas: {total_pastas_criadas}')

        if total_erros > 0:
            self.log('')
            self.log(f'❌ TOTAL DE ERROS: {total_erros}')
            self.log('')
            self.log('--- DETALHAMENTO DOS ERROS ---')
            for idx, (linha, erro) in enumerate(erros_detalhados, start=1):
                rc_e, po_e, forn_e = linha['rc'], linha['po'], linha['fornecedor']
                self.log(f'  {idx}. Fornecedor="{forn_e}" | RC="{rc_e}" | PO="{po_e}"')
                self.log(f'     Motivo: {erro}')
                self.log('')
        else:
            self.log('Nenhum erro durante a execucao.')

        if linhas_sem_nenhum_arquivo:
            self.log('')
            self.log(f'ℹ️ LINHAS SEM ARQUIVOS ENCONTRADOS ({len(linhas_sem_nenhum_arquivo)}):')
            self.log('  (pastas foram criadas, mas nenhum arquivo de proposta ou pedido foi localizado)')
            for rc_l, po_l, forn_l in linhas_sem_nenhum_arquivo:
                self.log(f'  • Fornecedor="{forn_l}" | RC="{rc_l}" | PO="{po_l}"')

        if linhas_sem_fornecedor > 0:
            self.log('')
            self.log(f'ℹ️ Linhas ignoradas (sem nome de fornecedor): {linhas_sem_fornecedor}')
